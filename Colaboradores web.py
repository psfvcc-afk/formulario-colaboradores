import streamlit as st
import pandas as pd
import dropbox
from datetime import datetime, date, timedelta
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
import calendar
import time
import json

st.set_page_config(
    page_title="Processamento Salarial v3.5",
    page_icon="💰",
    layout="wide"
)

# ==================== CONFIGURAÇÕES ====================

DROPBOX_APP_KEY = st.secrets["DROPBOX_APP_KEY"]
DROPBOX_APP_SECRET = st.secrets["DROPBOX_APP_SECRET"]
DROPBOX_REFRESH_TOKEN = st.secrets["DROPBOX_REFRESH_TOKEN"]
ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "adminpedro")

dbx = dropbox.Dropbox(
    app_key=DROPBOX_APP_KEY,
    app_secret=DROPBOX_APP_SECRET,
    oauth2_refresh_token=DROPBOX_REFRESH_TOKEN
)

EMPRESAS = {
    "Magnetic Sky Lda": {
        "path": "/Pedro Couto/Projectos/Alcalá_Arc_Amoreira/Gestão operacional/RH/Processamento Salários Magnetic/Gestão Colaboradores Magnetic.xlsx",
        "pasta_baixas": "/Pedro Couto/Projectos/Alcalá_Arc_Amoreira/Gestão operacional/RH/Baixas Médicas"
    },
    "CCM Retail Lda": {
        "path": "/Pedro Couto/Projectos/Pingo Doce/Pingo Doce/2. Operação/1. Recursos Humanos/Processamento salarial/Gestão Colaboradores.xlsx",
        "pasta_baixas": "/Pedro Couto/Projectos/Pingo Doce/Pingo Doce/2. Operação/1. Recursos Humanos/Baixas Médicas"
    }
}

# Categorias Profissionais por Empresa
CATEGORIAS_PROFISSIONAIS = {
    "Magnetic Sky Lda": [
        "Cozinheiro de 3ª",
        "Empregado de mesa/Balcão/Snack/Selfservice"
    ],
    "CCM Retail Lda": [
        "Gerente de Loja",
        "Chefe de Secção Especializado",
        "Op. Supermercado I",
        "Op. Supermercado Qualificado",
        "Cortador de Carnes Verdes I"
    ]
}

FERIADOS_NACIONAIS_2025 = [
    date(2025, 1, 1), date(2025, 4, 18), date(2025, 4, 20), date(2025, 4, 25),
    date(2025, 5, 1), date(2025, 6, 10), date(2025, 6, 19), date(2025, 8, 15),
    date(2025, 10, 5), date(2025, 11, 1), date(2025, 12, 1), date(2025, 12, 8),
    date(2025, 12, 25)
]

MOTIVOS_RESCISAO = [
    "Denúncia pela entidade patronal - período experimental",
    "Denúncia pelo trabalhador - período experimental",
    "Caducidade contrato a termo",
    "Denúncia pelo trabalhador - aviso prévio parcial",
    "Denúncia pelo trabalhador - aviso prévio completo",
    "Denúncia pelo trabalhador - sem aviso prévio",
    "Denúncia pela entidade patronal - excesso faltas",
    "Despedimento por facto imputável ao trabalhador",
    "Despedimento colectivo",
    "Despedimento por extinção do posto de trabalho",
    "Despedimento por inadaptação",
    "Revogação por acordo",
    "Reforma por velhice",
    "Reforma por invalidez",
    "Falecimento",
    "Outro (especificar em observações)"
]

COLUNAS_SNAPSHOT = [
    "Nome Completo", "Ano", "Mês", "Nº Horas/Semana", "Subsídio Alimentação Diário",
    "Número Pingo Doce", "Salário Bruto", "Vencimento Hora", 
    "Estado Civil", "Nº Titulares", "Nº Dependentes", "Deficiência",
    "IRS Percentagem Fixa", "IRS Modo Calculo",
    "Cartão Refeição", "Sub Férias Tipo", "Sub Natal Tipo",
    "Status", "Data Rescisão", "Motivo Rescisão", 
    "NIF", "NISS", "Data de Admissão", "IBAN", "Secção", 
    "E-mail", "Data de Nascimento", "Nº CC", "Validade CC", 
    "Nacionalidade", "Telemóvel", "Bairro Fiscal", "Morada", 
    "Cod Postal", "Categoria Profissional", "Timestamp"
]

COLUNAS_FALTAS_BAIXAS = [
    "Nome Completo", "Ano", "Mês", "Tipo", "Data Início", "Data Fim", 
    "Dias Úteis", "Dias Totais", "Observações", "Ficheiro Anexo", "Timestamp"
]

COLUNAS_HORAS_EXTRAS = [
    "Nome Completo", "Ano", "Mês", "Horas Noturnas", "Horas Domingos", 
    "Horas Feriados", "Horas Extra", "Outros Proveitos", "Observações", "Timestamp"
]

ESTADOS_CIVIS = ["Solteiro", "Casado Único Titular", "Casado Dois Titulares"]
HORAS_PERMITIDAS = [16, 20, 40]

MAPEAMENTO_ESTADO_CIVIL = {
    "Não Casado": "Solteiro",
    "Casado 1": "Casado Único Titular",
    "Casado 2": "Casado Dois Titulares",
    "Solteiro": "Solteiro",
    "Casado Único Titular": "Casado Único Titular",
    "Casado Dois Titulares": "Casado Dois Titulares"
}

MAPEAMENTO_TIPO_IRS = {
    "Automático (por tabela)": "Tabela",
    "Percentagem fixa": "Fixa",
    "Tabela": "Tabela",
    "Fixa": "Fixa",
    "Percentagem Fixa": "Fixa"
}

MAPEAMENTO_DEFICIENCIA = {
    "Sim": "Sim",
    "Não": "Não",
    "sim": "Sim",
    "não": "Não",
    "S": "Sim",
    "N": "Não"
}

# Categorias de campos para Output - v3.5 ATUALIZADO
CATEGORIAS_CAMPOS = {
    "👤 Dados Pessoais": [
        "Nome Completo", "Bairro Fiscal", "Categoria Profissional", "Cod Postal", 
        "Data de Admissão", "Data de Nascimento", "E-mail", "IBAN", 
        "Morada", "Nacionalidade", "NIF", "NISS", "Nº CC", 
        "Estado Civil", "Nº Titulares", "Nº Dependentes", "Deficiência",
        "Telemóvel", "Validade CC"
    ],
    "💼 Dados Profissionais": [
        "Status", "Secção", "Número Pingo Doce", "Data Rescisão", "Motivo Rescisão"
    ],
    "💰 Dados Salariais": [
        "Salário Bruto", "Nº Horas/Semana", "Vencimento Hora",
        "Subsídio Alimentação Diário", "Cartão Refeição",
        "Sub Férias Tipo", "Sub Natal Tipo"
    ],
    "📊 Dados IRS": [
        "IRS Modo Calculo", "IRS Percentagem Fixa"
    ],
    "🏖️ Faltas e Baixas": [
        "Total Faltas (dias)", "Total Baixas (dias)", "Total Faltas+Baixas"
    ],
    "⏰ Horas Extras": [
        "Horas Noturnas", "Horas Domingos", "Horas Feriados",
        "Horas Extra", "Total Horas Extras", "Outros Proveitos"
    ]
}

# ==================== SESSION STATE ====================

if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'salario_minimo' not in st.session_state:
    st.session_state.salario_minimo = 870.0
if 'feriados_municipais' not in st.session_state:
    st.session_state.feriados_municipais = [date(2025, 1, 14)]
if 'ultimo_reload' not in st.session_state:
    st.session_state.ultimo_reload = datetime.now()
if 'tabela_irs' not in st.session_state:
    st.session_state.tabela_irs = None
if 'dados_processamento' not in st.session_state:
    st.session_state.dados_processamento = {}
if 'empresa_selecionada' not in st.session_state:
    st.session_state.empresa_selecionada = list(EMPRESAS.keys())[0]
if 'mes_selecionado' not in st.session_state:
    st.session_state.mes_selecionado = datetime.now().month
if 'ano_selecionado' not in st.session_state:
    st.session_state.ano_selecionado = 2025
if 'colaborador_selecionado' not in st.session_state:
    st.session_state.colaborador_selecionado = None
if 'templates_relatorios' not in st.session_state:
    st.session_state.templates_relatorios = {}
if 'password_incorrect' not in st.session_state:
    st.session_state.password_incorrect = False
if 'campos_selecionados_output' not in st.session_state:
    st.session_state.campos_selecionados_output = []

# ==================== FUNÇÕES DE AUTENTICAÇÃO ====================

def check_password():
    def password_entered():
        if "password" in st.session_state:
            if st.session_state["password"] == ADMIN_PASSWORD:
                st.session_state.authenticated = True
                del st.session_state["password"]
            else:
                st.session_state.authenticated = False
                st.session_state.password_incorrect = True
    
    if not st.session_state.authenticated:
        st.title("🔒 Processamento Salarial - Login")
        st.markdown("---")
        st.text_input("Password de Administrador", type="password", key="password", on_change=password_entered)
        
        if st.session_state.get("password_incorrect", False):
            st.error("❌ Password incorreta")
            st.session_state.password_incorrect = False
        
        return False
    return True

# ==================== FUNÇÕES DE MAPEAMENTO ====================

def normalizar_estado_civil(valor):
    if pd.isna(valor) or valor == '':
        return "Solteiro"
    valor_str = str(valor).strip()
    return MAPEAMENTO_ESTADO_CIVIL.get(valor_str, "Solteiro")

def normalizar_tipo_irs(valor):
    if pd.isna(valor) or valor == '':
        return "Tabela"
    valor_str = str(valor).strip()
    return MAPEAMENTO_TIPO_IRS.get(valor_str, "Tabela")

def normalizar_deficiencia(valor):
    if pd.isna(valor) or valor == '':
        return "Não"
    valor_str = str(valor).strip()
    return MAPEAMENTO_DEFICIENCIA.get(valor_str, "Não")

def normalizar_percentagem_irs(valor):
    if pd.isna(valor) or valor == '':
        return 0.0
    try:
        return float(valor)
    except:
        return 0.0

def normalizar_sim_nao(valor):
    if pd.isna(valor) or valor == '':
        return "Não"
    valor_str = str(valor).strip()
    if valor_str in ["Sim", "sim", "S", "s", "Yes", "yes", "Y", "y"]:
        return "Sim"
    return "Não"

def normalizar_tipo_subsidio(valor):
    if pd.isna(valor) or valor == '':
        return "Duodécimos"
    valor_str = str(valor).strip()
    if valor_str in ["Total", "total", "T"]:
        return "Total"
    if valor_str in ["Não Pagar", "Nao Pagar", "Não", "Nao"]:
        return "Não Pagar"
    return "Duodécimos"

# ==================== FUNÇÕES DROPBOX ====================

def get_nome_aba_snapshot(ano, mes):
    return f"Estado_{ano}_{mes:02d}"

def get_nome_aba_faltas_baixas(ano, mes):
    return f"Faltas_Baixas_{ano}_{mes:02d}"

def get_nome_aba_horas_extras(ano, mes):
    return f"Extras_{ano}_{mes:02d}"

def criar_pasta_dropbox(path):
    """Cria pasta na Dropbox se não existir"""
    try:
        dbx.files_get_metadata(path)
        return True
    except:
        try:
            dbx.files_create_folder_v2(path)
            return True
        except Exception as e:
            st.error(f"Erro ao criar pasta: {e}")
            return False

def upload_ficheiro_baixa(empresa, ano, mes, colaborador, file):
    """Upload de ficheiro de baixa médica para Dropbox"""
    try:
        pasta_base = EMPRESAS[empresa]["pasta_baixas"]
        pasta_ano = f"{pasta_base}/{ano}"
        pasta_mes = f"{pasta_ano}/{mes:02d}_{calendar.month_name[mes]}"
        
        criar_pasta_dropbox(pasta_base)
        criar_pasta_dropbox(pasta_ano)
        criar_pasta_dropbox(pasta_mes)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_limpo = colaborador.replace(" ", "_")
        extensao = file.name.split(".")[-1]
        nome_ficheiro = f"{nome_limpo}_{timestamp}.{extensao}"
        
        caminho_completo = f"{pasta_mes}/{nome_ficheiro}"
        
        file.seek(0)
        dbx.files_upload(file.read(), caminho_completo, mode=dropbox.files.WriteMode.overwrite)
        
        return caminho_completo
        
    except Exception as e:
        st.error(f"❌ Erro ao fazer upload: {e}")
        return None

def download_excel(empresa):
    try:
        file_path = EMPRESAS[empresa]["path"]
        _, response = dbx.files_download(file_path)
        return BytesIO(response.content)
    except Exception as e:
        st.error(f"❌ Erro ao baixar Excel: {e}")
        return None

def criar_backup_dropbox(empresa):
    """Cria backup do ficheiro antes de modificar - v3.5"""
    try:
        file_path = EMPRESAS[empresa]["path"]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.replace(".xlsx", f"_backup_{timestamp}.xlsx")
        
        dbx.files_copy_v2(file_path, backup_path)
        st.info(f"💾 Backup criado: ...{backup_path[-40:]}")
        return True
    except Exception as e:
        st.warning(f"⚠️ Não foi possível criar backup: {e}")
        return False

def garantir_aba(wb, nome_aba, colunas):
    """Garante que a aba existe, criando-a se necessário - v3.5"""
    if nome_aba not in wb.sheetnames:
        ws = wb.create_sheet(nome_aba)
        ws.append(colunas)
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return True
    return False

def validar_workbook(wb):
    """Valida integridade do workbook - v3.5"""
    try:
        if "Colaboradores" not in wb.sheetnames:
            st.error("🚨 CRÍTICO: Aba 'Colaboradores' não encontrada!")
            return False
        
        ws_colab = wb["Colaboradores"]
        if ws_colab.max_row < 2:
            st.error("🚨 CRÍTICO: Aba 'Colaboradores' está vazia!")
            return False
        
        # Verificar se há pelo menos uma coluna
        if ws_colab.max_column < 1:
            st.error("🚨 CRÍTICO: Aba 'Colaboradores' sem colunas!")
            return False
        
        return True
    except Exception as e:
        st.error(f"🚨 Erro ao validar workbook: {e}")
        return False

def upload_excel_seguro(empresa, wb):
    """Upload com validação robusta e backup - v3.5 MELHORADO"""
    try:
        # VALIDAÇÃO PRÉ-UPLOAD
        if not validar_workbook(wb):
            st.error("❌ Validação falhou - upload cancelado!")
            return False
        
        # CRIAR BACKUP
        criar_backup_dropbox(empresa)
        
        # PREPARAR UPLOAD
        file_path = EMPRESAS[empresa]["path"]
        output = BytesIO()
        
        # SALVAR com tratamento de erro
        try:
            wb.save(output)
        except Exception as save_error:
            st.error(f"❌ Erro ao salvar workbook: {save_error}")
            return False
        
        output.seek(0)
        
        # VALIDAR tamanho do ficheiro
        file_size = output.getbuffer().nbytes
        if file_size < 1000:  # Menos de 1KB é suspeito
            st.error(f"❌ Ficheiro muito pequeno ({file_size} bytes) - upload cancelado!")
            return False
        
        # UPLOAD
        dbx.files_upload(output.read(), file_path, mode=dropbox.files.WriteMode.overwrite)
        
        # VERIFICAR upload
        time.sleep(0.5)  # Pequena pausa para garantir sincronização
        
        ws_colab = wb["Colaboradores"]
        st.success(f"✅ Excel salvo ({ws_colab.max_row-1} colaboradores, {file_size:,} bytes)")
        
        return True
        
    except Exception as e:
        st.error(f"❌ Erro ao enviar Excel: {e}")
        import traceback
        st.error(f"🔍 Detalhes: {traceback.format_exc()}")
        return False

# ==================== FUNÇÕES DE CÁLCULO ====================

def calcular_vencimento_hora(salario_bruto, horas_semana):
    if horas_semana == 0:
        return 0
    return (salario_bruto * 12) / (52 * horas_semana)

def calcular_vencimento_ajustado(salario_bruto, dias_faltas, dias_baixas):
    dias_pagos = 30 - dias_faltas - dias_baixas
    dias_pagos = max(dias_pagos, 0)
    return (salario_bruto / 30) * dias_pagos

def calcular_dias_entre_datas(data_inicio, data_fim, feriados_list):
    """Calcula dias úteis e totais entre duas datas"""
    if data_inicio > data_fim:
        return 0, 0
    
    dias_totais = (data_fim - data_inicio).days + 1
    dias_uteis = 0
    
    data_atual = data_inicio
    while data_atual <= data_fim:
        if data_atual.weekday() < 5 and data_atual not in feriados_list:
            dias_uteis += 1
        data_atual += timedelta(days=1)
    
    return dias_uteis, dias_totais

def calcular_dias_uteis(ano, mes, feriados_list):
    num_dias = calendar.monthrange(ano, mes)[1]
    dias_uteis = 0
    for dia in range(1, num_dias + 1):
        data = date(ano, mes, dia)
        if data.weekday() < 5 and data not in feriados_list:
            dias_uteis += 1
    return dias_uteis

def carregar_tabela_irs_excel(uploaded_file):
    try:
        xls = pd.ExcelFile(uploaded_file)
        st.success(f"✅ Ficheiro carregado! Abas: {', '.join(xls.sheet_names)}")
        st.session_state.tabela_irs = xls
        return xls
    except Exception as e:
        st.error(f"❌ Erro ao carregar tabela: {e}")
        return None

def calcular_irs_por_tabela(base_incidencia, estado_civil, num_dependentes, tem_deficiencia=False):
    reducao_dependentes = num_dependentes * 0.01
    
    if base_incidencia <= 820:
        taxa = 0.135
    elif base_incidencia <= 1200:
        taxa = 0.18
    elif base_incidencia <= 1700:
        taxa = 0.23
    elif base_incidencia <= 2500:
        taxa = 0.265
    else:
        taxa = 0.32
    
    taxa_final = max(taxa - reducao_dependentes, 0.05)
    
    if estado_civil == "Casado Único Titular":
        taxa_final *= 0.85
    
    return base_incidencia * taxa_final

def calcular_irs(base_incidencia, modo_calculo, percentagem_fixa, estado_civil, num_dependentes, tem_deficiencia=False):
    if modo_calculo == "Fixa":
        taxa = percentagem_fixa / 100
        irs = base_incidencia * taxa
        return irs
    else:
        return calcular_irs_por_tabela(base_incidencia, estado_civil, num_dependentes, tem_deficiencia)

# ==================== FUNÇÕES DE DADOS BASE ====================

def carregar_dados_base(empresa):
    """Lê sempre da aba 'Colaboradores' - v3.5"""
    excel_file = download_excel(empresa)
    if excel_file:
        try:
            df = pd.read_excel(excel_file, sheet_name="Colaboradores")
            
            # Garantir colunas essenciais
            colunas_essenciais = {
                'Status': 'Ativo',
                'Salário Bruto': 870.0,
                'Sub Férias Tipo': 'Duodécimos',
                'Sub Natal Tipo': 'Duodécimos',
                'Cartão Refeição': 'Não',
                'Data Rescisão': '',
                'Motivo Rescisão': '',
                'IBAN': '',
                'E-mail': '',
                'Data de Nascimento': '',
                'Nº CC': '',
                'Validade CC': '',
                'Nacionalidade': '',
                'Telemóvel': '',
                'Bairro Fiscal': '',
                'Morada': '',
                'Cod Postal': '',
                'Categoria Profissional': ''
            }
            
            for col, default in colunas_essenciais.items():
                if col not in df.columns:
                    df[col] = default
            
            # Normalizar Status
            df.loc[df['Status'].isna() | (df['Status'] == ''), 'Status'] = 'Ativo'
            
            return df
        except Exception as e:
            st.error(f"❌ Erro ao ler aba Colaboradores: {e}")
    return pd.DataFrame()

def carregar_colaboradores_ativos(empresa, ano=None, mes=None):
    """Lê da aba 'Colaboradores' onde Status = 'Ativo'"""
    df_base = carregar_dados_base(empresa)
    
    if df_base.empty:
        return []
    
    if 'Status' in df_base.columns:
        df_ativos = df_base[df_base['Status'] == 'Ativo']
    else:
        df_ativos = df_base
    
    colaboradores = df_ativos['Nome Completo'].tolist()
    
    return colaboradores

def atualizar_status_colaborador(empresa, colaborador, novo_status):
    """Atualiza Status APENAS na aba Colaboradores - v3.5"""
    try:
        excel_file = download_excel(empresa)
        if not excel_file:
            return False
        
        df = pd.read_excel(excel_file, sheet_name="Colaboradores")
        
        if 'Status' not in df.columns:
            df['Status'] = 'Ativo'
        
        mask = df['Nome Completo'] == colaborador
        if mask.any():
            df.loc[mask, 'Status'] = novo_status
        else:
            st.error(f"❌ Colaborador '{colaborador}' não encontrado")
            return False
        
        wb = load_workbook(excel_file, data_only=False)
        
        if "Colaboradores" in wb.sheetnames:
            idx = wb.sheetnames.index("Colaboradores")
            del wb["Colaboradores"]
            ws = wb.create_sheet("Colaboradores", idx)
        else:
            ws = wb.create_sheet("Colaboradores", 0)
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        if upload_excel_seguro(empresa, wb):
            st.success(f"✅ Status de '{colaborador}' → '{novo_status}'")
            return True
        
        return False
        
    except Exception as e:
        st.error(f"❌ Erro: {e}")
        return False

def criar_snapshot_inicial(empresa, colaborador, ano, mes):
    """Cria snapshot inicial - v3.5 com novos campos"""
    df_base = carregar_dados_base(empresa)
    dados_colab = df_base[df_base['Nome Completo'] == colaborador]
    
    if dados_colab.empty:
        return None
    
    dados = dados_colab.iloc[0]
    horas_semana = float(dados.get('Nº Horas/Semana', 40))
    salario_bruto = float(dados.get('Salário Bruto', 870.0))
    
    estado_civil_raw = dados.get('Estado Civil', 'Solteiro')
    estado_civil = normalizar_estado_civil(estado_civil_raw)
    
    tipo_irs_raw = dados.get('Tipo IRS', 'Tabela')
    tipo_irs = normalizar_tipo_irs(tipo_irs_raw)
    
    perc_irs_raw = dados.get('% IRS Fixa', 0)
    perc_irs = normalizar_percentagem_irs(perc_irs_raw)
    
    deficiencia_raw = dados.get('Pessoa com Deficiência', 'Não')
    deficiencia = normalizar_deficiencia(deficiencia_raw)
    
    cartao_refeicao = normalizar_sim_nao(dados.get('Cartão Refeição', 'Não'))
    sub_ferias_tipo = normalizar_tipo_subsidio(dados.get('Sub Férias Tipo', 'Duodécimos'))
    sub_natal_tipo = normalizar_tipo_subsidio(dados.get('Sub Natal Tipo', 'Duodécimos'))
    
    status = dados.get('Status', 'Ativo')
    
    snapshot = {
        "Nome Completo": colaborador,
        "Ano": ano,
        "Mês": mes,
        "Nº Horas/Semana": horas_semana,
        "Subsídio Alimentação Diário": float(dados.get('Subsídio Alimentação Diário', 5.96)),
        "Número Pingo Doce": str(dados.get('Número Pingo Doce', '')),
        "Salário Bruto": salario_bruto,
        "Vencimento Hora": calcular_vencimento_hora(salario_bruto, horas_semana),
        "Estado Civil": estado_civil,
        "Nº Titulares": int(dados.get('Nº Titulares', 2)),
        "Nº Dependentes": int(dados.get('Nº Dependentes', 0)),
        "Deficiência": deficiencia,
        "IRS Percentagem Fixa": perc_irs,
        "IRS Modo Calculo": tipo_irs,
        "Cartão Refeição": cartao_refeicao,
        "Sub Férias Tipo": sub_ferias_tipo,
        "Sub Natal Tipo": sub_natal_tipo,
        "Status": status,
        "Data Rescisão": str(dados.get('Data Rescisão', '')),
        "Motivo Rescisão": str(dados.get('Motivo Rescisão', '')),
        "NIF": str(dados.get('NIF', '')),
        "NISS": str(dados.get('NISS', '')),
        "Data de Admissão": str(dados.get('Data de Admissão', '')),
        "IBAN": str(dados.get('IBAN', '')),
        "Secção": str(dados.get('Secção', '')),
        "E-mail": str(dados.get('E-mail', '')),
        "Data de Nascimento": str(dados.get('Data de Nascimento', '')),
        "Nº CC": str(dados.get('Nº CC', '')),
        "Validade CC": str(dados.get('Validade CC', '')),
        "Nacionalidade": str(dados.get('Nacionalidade', '')),
        "Telemóvel": str(dados.get('Telemóvel', '')),
        "Bairro Fiscal": str(dados.get('Bairro Fiscal', '')),
        "Morada": str(dados.get('Morada', '')),
        "Cod Postal": str(dados.get('Cod Postal', '')),
        "Categoria Profissional": str(dados.get('Categoria Profissional', '')),
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    return snapshot

def carregar_ultimo_snapshot(empresa, colaborador, ano, mes):
    """Carrega último snapshot com dados ATUALIZADOS - v3.5"""
    excel_file = download_excel(empresa)
    if not excel_file:
        return None
    
    try:
        wb = load_workbook(excel_file, data_only=False)
        nome_aba = get_nome_aba_snapshot(ano, mes)
        
        if nome_aba in wb.sheetnames:
            df = pd.read_excel(excel_file, sheet_name=nome_aba)
            df_colab = df[df['Nome Completo'] == colaborador]
            
            if not df_colab.empty:
                snapshot = df_colab.iloc[-1].to_dict()
                
                # Atualizar com dados mais recentes da aba Colaboradores
                df_base = carregar_dados_base(empresa)
                dados_colab = df_base[df_base['Nome Completo'] == colaborador]
                
                if not dados_colab.empty:
                    dados = dados_colab.iloc[0]
                    
                    # Campos dinâmicos
                    snapshot['Nº Horas/Semana'] = float(dados.get('Nº Horas/Semana', snapshot.get('Nº Horas/Semana', 40)))
                    snapshot['Subsídio Alimentação Diário'] = float(dados.get('Subsídio Alimentação Diário', snapshot.get('Subsídio Alimentação Diário', 5.96)))
                    snapshot['Número Pingo Doce'] = str(dados.get('Número Pingo Doce', snapshot.get('Número Pingo Doce', '')))
                    snapshot['Salário Bruto'] = float(dados.get('Salário Bruto', snapshot.get('Salário Bruto', 870.0)))
                    
                    horas = float(snapshot['Nº Horas/Semana'])
                    snapshot['Vencimento Hora'] = calcular_vencimento_hora(snapshot['Salário Bruto'], horas)
                    
                    snapshot['Cartão Refeição'] = normalizar_sim_nao(dados.get('Cartão Refeição', snapshot.get('Cartão Refeição', 'Não')))
                    snapshot['Sub Férias Tipo'] = normalizar_tipo_subsidio(dados.get('Sub Férias Tipo', snapshot.get('Sub Férias Tipo', 'Duodécimos')))
                    snapshot['Sub Natal Tipo'] = normalizar_tipo_subsidio(dados.get('Sub Natal Tipo', snapshot.get('Sub Natal Tipo', 'Duodécimos')))
                    
                    snapshot['Estado Civil'] = normalizar_estado_civil(dados.get('Estado Civil', snapshot.get('Estado Civil', 'Solteiro')))
                    snapshot['Nº Titulares'] = int(dados.get('Nº Titulares', snapshot.get('Nº Titulares', 2)))
                    snapshot['Nº Dependentes'] = int(dados.get('Nº Dependentes', snapshot.get('Nº Dependentes', 0)))
                    snapshot['Deficiência'] = normalizar_deficiencia(dados.get('Pessoa com Deficiência', snapshot.get('Deficiência', 'Não')))
                    snapshot['IRS Modo Calculo'] = normalizar_tipo_irs(dados.get('Tipo IRS', snapshot.get('IRS Modo Calculo', 'Tabela')))
                    snapshot['IRS Percentagem Fixa'] = normalizar_percentagem_irs(dados.get('% IRS Fixa', snapshot.get('IRS Percentagem Fixa', 0)))
                    snapshot['IBAN'] = str(dados.get('IBAN', snapshot.get('IBAN', '')))
                    
                    # Novos campos v3.5
                    snapshot['E-mail'] = str(dados.get('E-mail', snapshot.get('E-mail', '')))
                    snapshot['Data de Nascimento'] = str(dados.get('Data de Nascimento', snapshot.get('Data de Nascimento', '')))
                    snapshot['Nº CC'] = str(dados.get('Nº CC', snapshot.get('Nº CC', '')))
                    snapshot['Validade CC'] = str(dados.get('Validade CC', snapshot.get('Validade CC', '')))
                    snapshot['Nacionalidade'] = str(dados.get('Nacionalidade', snapshot.get('Nacionalidade', '')))
                    snapshot['Telemóvel'] = str(dados.get('Telemóvel', snapshot.get('Telemóvel', '')))
                    snapshot['Bairro Fiscal'] = str(dados.get('Bairro Fiscal', snapshot.get('Bairro Fiscal', '')))
                    snapshot['Morada'] = str(dados.get('Morada', snapshot.get('Morada', '')))
                    snapshot['Cod Postal'] = str(dados.get('Cod Postal', snapshot.get('Cod Postal', '')))
                    snapshot['Categoria Profissional'] = str(dados.get('Categoria Profissional', snapshot.get('Categoria Profissional', '')))
                    
                    snapshot['Data Rescisão'] = str(dados.get('Data Rescisão', snapshot.get('Data Rescisão', '')))
                    snapshot['Motivo Rescisão'] = str(dados.get('Motivo Rescisão', snapshot.get('Motivo Rescisão', '')))
                
                if 'Status' not in snapshot or pd.isna(snapshot['Status']) or snapshot['Status'] == '':
                    snapshot['Status'] = 'Ativo'
                
                st.caption(f"📸 Snapshot {ano}-{mes:02d}: {snapshot.get('Timestamp', 'N/A')}")
                return snapshot
        
        snapshot = criar_snapshot_inicial(empresa, colaborador, ano, mes)
        if snapshot:
            st.caption(f"📸 Criado da aba Colaboradores")
        return snapshot
        
    except Exception as e:
        st.error(f"❌ Erro: {e}")
        return None

def gravar_snapshot(empresa, snapshot):
    """Grava snapshot SEM mexer na aba Colaboradores - v3.5"""
    try:
        if 'Status' not in snapshot or pd.isna(snapshot['Status']) or snapshot['Status'] == '':
            snapshot['Status'] = 'Ativo'
        
        ano = snapshot['Ano']
        mes = snapshot['Mês']
        nome_aba = get_nome_aba_snapshot(ano, mes)
        
        excel_file = download_excel(empresa)
        if not excel_file:
            return False
        
        wb = load_workbook(excel_file, data_only=False)
        
        if not validar_workbook(wb):
            return False
        
        aba_criada = garantir_aba(wb, nome_aba, COLUNAS_SNAPSHOT)
        
        ws = wb[nome_aba]
        
        nova_linha = []
        for col in COLUNAS_SNAPSHOT:
            valor = snapshot.get(col, '')
            if isinstance(valor, (int, float)):
                nova_linha.append(valor)
            else:
                nova_linha.append(str(valor) if valor else '')
        
        ws.append(nova_linha)
        
        sucesso = upload_excel_seguro(empresa, wb)
        
        if sucesso:
            linha = ws.max_row
            st.success(f"✅ Snapshot gravado (linha {linha})")
            if aba_criada:
                st.success(f"✨ Aba '{nome_aba}' foi criada")
            return True
        
        return False
        
    except Exception as e:
        st.error(f"❌ Erro ao gravar: {e}")
        import traceback
        st.error(f"🔍 Detalhes: {traceback.format_exc()}")
        return False

def gravar_falta_baixa(empresa, ano, mes, colaborador, tipo, data_inicio, data_fim, obs, ficheiro_path=None):
    """Grava registo de falta ou baixa - v3.5 ULTRA ROBUSTO"""
    try:
        excel_file = download_excel(empresa)
        if not excel_file:
            st.error("❌ Erro ao baixar Excel")
            return False
        
        # CARREGAR workbook com cuidado
        wb = load_workbook(excel_file, data_only=False)
        
        # VALIDAR antes de qualquer operação
        if not validar_workbook(wb):
            st.error("❌ Validação inicial falhou")
            return False
        
        nome_aba = get_nome_aba_faltas_baixas(ano, mes)
        
        # GARANTIR aba existe
        aba_foi_criada = garantir_aba(wb, nome_aba, COLUNAS_FALTAS_BAIXAS)
        
        if nome_aba not in wb.sheetnames:
            st.error(f"❌ ERRO CRÍTICO: Falha ao criar aba '{nome_aba}'")
            return False
        
        ws = wb[nome_aba]
        
        # CALCULAR dias
        feriados = FERIADOS_NACIONAIS_2025 + st.session_state.feriados_municipais
        dias_uteis, dias_totais = calcular_dias_entre_datas(data_inicio, data_fim, feriados)
        
        # PREPARAR linha
        nova_linha = [
            colaborador,
            ano,
            mes,
            tipo,
            data_inicio.strftime("%Y-%m-%d"),
            data_fim.strftime("%Y-%m-%d"),
            dias_uteis,
            dias_totais,
            obs,
            ficheiro_path if ficheiro_path else "",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        # ADICIONAR linha
        ws.append(nova_linha)
        
        # LOGS
        st.info(f"📊 Aba '{nome_aba}' agora tem {ws.max_row} linhas (incluindo cabeçalho)")
        
        # VALIDAR novamente antes de upload
        if not validar_workbook(wb):
            st.error("❌ Validação pós-modificação falhou")
            return False
        
        # UPLOAD SEGURO
        if upload_excel_seguro(empresa, wb):
            st.success(f"✅ {tipo} registada: {dias_uteis} dias úteis / {dias_totais} dias totais")
            if aba_foi_criada:
                st.success(f"✨ Aba '{nome_aba}' foi criada")
            return True
        
        return False
        
    except Exception as e:
        st.error(f"❌ Erro ao gravar falta/baixa: {e}")
        import traceback
        st.error(f"🔍 Detalhes: {traceback.format_exc()}")
        return False

def eliminar_registo_falta_baixa(empresa, ano, mes, linha_idx):
    """Elimina um registo específico de falta/baixa - v3.5"""
    try:
        excel_file = download_excel(empresa)
        if not excel_file:
            return False
        
        wb = load_workbook(excel_file, data_only=False)
        
        if not validar_workbook(wb):
            return False
        
        nome_aba = get_nome_aba_faltas_baixas(ano, mes)
        
        if nome_aba not in wb.sheetnames:
            st.error(f"❌ Aba '{nome_aba}' não encontrada!")
            return False
        
        ws = wb[nome_aba]
        linha_excel = linha_idx + 2
        
        if linha_excel > ws.max_row:
            st.error("❌ Linha inválida!")
            return False
        
        ws.delete_rows(linha_excel)
        
        if not validar_workbook(wb):
            return False
        
        if upload_excel_seguro(empresa, wb):
            st.success("✅ Registo eliminado!")
            return True
        
        return False
        
    except Exception as e:
        st.error(f"❌ Erro ao eliminar: {e}")
        return False

def gravar_horas_extras(empresa, ano, mes, colaborador, h_noturnas, h_domingos, h_feriados, h_extra, outros_prov, obs):
    """Grava registo de horas extras - v3.5 ULTRA ROBUSTO"""
    try:
        excel_file = download_excel(empresa)
        if not excel_file:
            st.error("❌ Erro ao baixar Excel")
            return False
        
        # CARREGAR workbook com cuidado
        wb = load_workbook(excel_file, data_only=False)
        
        # VALIDAR antes de qualquer operação
        if not validar_workbook(wb):
            st.error("❌ Validação inicial falhou")
            return False
        
        nome_aba = get_nome_aba_horas_extras(ano, mes)
        
        # GARANTIR aba existe
        aba_foi_criada = garantir_aba(wb, nome_aba, COLUNAS_HORAS_EXTRAS)
        
        if nome_aba not in wb.sheetnames:
            st.error(f"❌ ERRO CRÍTICO: Falha ao criar aba '{nome_aba}'")
            return False
        
        ws = wb[nome_aba]
        
        # PREPARAR linha
        nova_linha = [
            colaborador,
            ano,
            mes,
            h_noturnas,
            h_domingos,
            h_feriados,
            h_extra,
            outros_prov,
            obs,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        # ADICIONAR linha
        ws.append(nova_linha)
        
        # LOGS
        st.info(f"📊 Aba '{nome_aba}' agora tem {ws.max_row} linhas (incluindo cabeçalho)")
        
        # VALIDAR novamente antes de upload
        if not validar_workbook(wb):
            st.error("❌ Validação pós-modificação falhou")
            return False
        
        # UPLOAD SEGURO
        if upload_excel_seguro(empresa, wb):
            st.success(f"✅ Horas extras/proveitos registados")
            if aba_foi_criada:
                st.success(f"✨ Aba '{nome_aba}' foi criada")
            return True
        
        return False
        
    except Exception as e:
        st.error(f"❌ Erro ao gravar horas extras: {e}")
        import traceback
        st.error(f"🔍 Detalhes: {traceback.format_exc()}")
        return False

def eliminar_registo_horas_extras(empresa, ano, mes, linha_idx):
    """Elimina um registo específico de horas extras - v3.5"""
    try:
        excel_file = download_excel(empresa)
        if not excel_file:
            return False
        
        wb = load_workbook(excel_file, data_only=False)
        
        if not validar_workbook(wb):
            return False
        
        nome_aba = get_nome_aba_horas_extras(ano, mes)
        
        if nome_aba not in wb.sheetnames:
            st.error(f"❌ Aba '{nome_aba}' não encontrada!")
            return False
        
        ws = wb[nome_aba]
        linha_excel = linha_idx + 2
        
        if linha_excel > ws.max_row:
            st.error("❌ Linha inválida!")
            return False
        
        ws.delete_rows(linha_excel)
        
        if not validar_workbook(wb):
            return False
        
        if upload_excel_seguro(empresa, wb):
            st.success("✅ Registo eliminado!")
            return True
        
        return False
        
    except Exception as e:
        st.error(f"❌ Erro ao eliminar: {e}")
        return False

def carregar_faltas_baixas(empresa, ano, mes, colaborador=None):
    """Carrega faltas e baixas do mês"""
    try:
        excel_file = download_excel(empresa)
        if not excel_file:
            return pd.DataFrame()
        
        nome_aba = get_nome_aba_faltas_baixas(ano, mes)
        
        try:
            df = pd.read_excel(excel_file, sheet_name=nome_aba)
            
            if colaborador:
                df = df[df['Nome Completo'] == colaborador]
            
            return df
        except:
            return pd.DataFrame()
            
    except Exception as e:
        st.error(f"❌ Erro: {e}")
        return pd.DataFrame()

def carregar_horas_extras(empresa, ano, mes, colaborador=None):
    """Carrega horas extras do mês"""
    try:
        excel_file = download_excel(empresa)
        if not excel_file:
            return pd.DataFrame()
        
        nome_aba = get_nome_aba_horas_extras(ano, mes)
        
        try:
            df = pd.read_excel(excel_file, sheet_name=nome_aba)
            
            if colaborador:
                df = df[df['Nome Completo'] == colaborador]
            
            return df
        except:
            return pd.DataFrame()
            
    except Exception as e:
        st.error(f"❌ Erro: {e}")
        return pd.DataFrame()

def registar_rescisao_colaborador(empresa, colaborador, data_rescisao, motivo, obs):
    """Registra rescisão na aba Colaboradores - v3.5"""
    try:
        excel_file = download_excel(empresa)
        if not excel_file:
            return False
        
        df = pd.read_excel(excel_file, sheet_name="Colaboradores")
        
        mask = df['Nome Completo'] == colaborador
        if mask.any():
            df.loc[mask, 'Data Rescisão'] = data_rescisao.strftime("%Y-%m-%d")
            obs_completa = f"{motivo}"
            if obs:
                obs_completa += f" | Obs: {obs}"
            df.loc[mask, 'Motivo Rescisão'] = obs_completa
        else:
            st.error(f"❌ Colaborador '{colaborador}' não encontrado")
            return False
        
        wb = load_workbook(excel_file, data_only=False)
        
        if not validar_workbook(wb):
            return False
        
        if "Colaboradores" in wb.sheetnames:
            idx = wb.sheetnames.index("Colaboradores")
            del wb["Colaboradores"]
            ws = wb.create_sheet("Colaboradores", idx)
        else:
            ws = wb.create_sheet("Colaboradores", 0)
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        if upload_excel_seguro(empresa, wb):
            st.success(f"✅ Rescisão registada para '{colaborador}'")
            return True
        
        return False
        
    except Exception as e:
        st.error(f"❌ Erro: {e}")
        return False

def processar_calculo_salario(dados_form):
    salario_bruto = dados_form['salario_bruto']
    horas_semana = dados_form['horas_semana']
    sub_alimentacao_dia = dados_form['subsidio_alimentacao']
    vencimento_hora = calcular_vencimento_hora(salario_bruto, horas_semana)
    
    dias_faltas = dados_form['dias_faltas']
    dias_baixas = dados_form['dias_baixas']
    dias_uteis_trabalhados = dados_form['dias_uteis_trabalhados']
    
    horas_noturnas = dados_form.get('horas_noturnas', 0)
    horas_domingos = dados_form.get('horas_domingos', 0)
    horas_feriados = dados_form.get('horas_feriados', 0)
    horas_extra = dados_form.get('horas_extra', 0)
    
    vencimento_ajustado = calcular_vencimento_ajustado(salario_bruto, dias_faltas, dias_baixas)
    
    sub_alimentacao = sub_alimentacao_dia * dias_uteis_trabalhados
    trabalho_noturno = horas_noturnas * vencimento_hora * 0.25
    domingos = horas_domingos * vencimento_hora
    feriados = horas_feriados * vencimento_hora * 2
    
    sub_ferias_tipo = dados_form.get('sub_ferias_tipo', 'Duodécimos')
    if sub_ferias_tipo == 'Total':
        sub_ferias = salario_bruto
    elif sub_ferias_tipo == 'Não Pagar':
        sub_ferias = 0
    else:
        sub_ferias = salario_bruto / 12
    
    sub_natal_tipo = dados_form.get('sub_natal_tipo', 'Duodécimos')
    if sub_natal_tipo == 'Total':
        sub_natal = salario_bruto
    elif sub_natal_tipo == 'Não Pagar':
        sub_natal = 0
    else:
        sub_natal = salario_bruto / 12
    
    banco_horas_valor = vencimento_hora * horas_extra
    outros_proveitos = dados_form.get('outros_proveitos', 0)
    
    total_remuneracoes = (vencimento_ajustado + sub_alimentacao + trabalho_noturno + 
                          domingos + feriados + sub_ferias + sub_natal + 
                          banco_horas_valor + outros_proveitos)
    
    base_ss = total_remuneracoes - sub_alimentacao
    seg_social = base_ss * 0.11
    
    base_irs = (vencimento_ajustado + trabalho_noturno + domingos + feriados + 
                sub_ferias + sub_natal + banco_horas_valor + outros_proveitos)
    
    irs = calcular_irs(
        base_incidencia=base_irs,
        modo_calculo=dados_form.get('irs_modo', 'Tabela'),
        percentagem_fixa=dados_form.get('irs_percentagem_fixa', 0),
        estado_civil=dados_form.get('estado_civil', 'Solteiro'),
        num_dependentes=dados_form.get('num_dependentes', 0),
        tem_deficiencia=dados_form.get('tem_deficiencia', False)
    )
    
    desconto_especie = 0
    cartao_refeicao = dados_form.get('cartao_refeicao', False)
    
    if cartao_refeicao:
        desconto_especie = sub_alimentacao
    
    total_descontos = seg_social + irs + desconto_especie
    
    liquido = total_remuneracoes - total_descontos
    
    return {
        'vencimento_hora': vencimento_hora,
        'vencimento_ajustado': vencimento_ajustado,
        'sub_alimentacao': sub_alimentacao,
        'trabalho_noturno': trabalho_noturno,
        'domingos': domingos,
        'feriados': feriados,
        'sub_ferias': sub_ferias,
        'sub_natal': sub_natal,
        'banco_horas_valor': banco_horas_valor,
        'outros_proveitos': outros_proveitos,
        'total_remuneracoes': total_remuneracoes,
        'base_ss': base_ss,
        'seg_social': seg_social,
        'base_irs': base_irs,
        'irs': irs,
        'desconto_especie': desconto_especie,
        'cartao_refeicao': cartao_refeicao,
        'total_descontos': total_descontos,
        'liquido': liquido
    }

def calcular_ftes_e_estatisticas(empresa, ano=None, mes=None):
    """Calcula FTEs e estatísticas por secção"""
    df_base = carregar_dados_base(empresa)
    
    if df_base.empty:
        return None
    
    df_ativos = df_base[df_base['Status'] == 'Ativo'].copy()
    
    if df_ativos.empty:
        return None
    
    if 'Secção' not in df_ativos.columns:
        df_ativos['Secção'] = 'Sem Secção'
    
    df_ativos['Secção'] = df_ativos['Secção'].fillna('Sem Secção')
    df_ativos['Secção'] = df_ativos['Secção'].replace('', 'Sem Secção')
    
    estatisticas = []
    
    for seccao in df_ativos['Secção'].unique():
        df_seccao = df_ativos[df_ativos['Secção'] == seccao]
        
        num_colaboradores = len(df_seccao)
        horas_totais = df_seccao['Nº Horas/Semana'].sum()
        ftes = horas_totais / 40
        
        h16 = len(df_seccao[df_seccao['Nº Horas/Semana'] == 16])
        h20 = len(df_seccao[df_seccao['Nº Horas/Semana'] == 20])
        h40 = len(df_seccao[df_seccao['Nº Horas/Semana'] == 40])
        
        estatisticas.append({
            'Secção': seccao,
            'Nº Colaboradores': num_colaboradores,
            '16h': h16,
            '20h': h20,
            '40h': h40,
            'Total Horas/Semana': horas_totais,
            'FTEs': round(ftes, 2)
        })
    
    df_stats = pd.DataFrame(estatisticas)
    df_stats = df_stats.sort_values('Secção')
    
    total_colaboradores = df_ativos.shape[0]
    total_horas = df_ativos['Nº Horas/Semana'].sum()
    total_ftes = total_horas / 40
    
    return {
        'df_stats': df_stats,
        'total_colaboradores': total_colaboradores,
        'total_horas': total_horas,
        'total_ftes': round(total_ftes, 2),
        'df_ativos': df_ativos
    }

def carregar_dados_completos_relatorio(empresa, ano, mes, filtros):
    """Carrega dados completos incluindo faltas/baixas e horas extras - v3.5"""
    try:
        df_base = carregar_dados_base(empresa)
        
        if df_base.empty:
            return None
        
        df_filtrado = df_base.copy()
        
        # Filtro de Status
        if filtros.get('status') and filtros['status'] != 'Todos':
            df_filtrado = df_filtrado[df_filtrado['Status'] == filtros['status']]
        
        # Filtro de Secção
        if filtros.get('seccao') and filtros['seccao'] != 'Todas':
            df_filtrado = df_filtrado[df_filtrado['Secção'] == filtros['seccao']]
        
        # v3.5: Filtro de Colaboradores
        if filtros.get('colaboradores') and len(filtros['colaboradores']) > 0:
            df_filtrado = df_filtrado[df_filtrado['Nome Completo'].isin(filtros['colaboradores'])]
        
        df_faltas_baixas = carregar_faltas_baixas(empresa, ano, mes)
        df_horas_extras = carregar_horas_extras(empresa, ano, mes)
        
        resultado = []
        
        for _, row in df_filtrado.iterrows():
            nome = row['Nome Completo']
            dados_colab = row.to_dict()
            
            # Faltas e Baixas
            if not df_faltas_baixas.empty:
                df_colab_faltas = df_faltas_baixas[df_faltas_baixas['Nome Completo'] == nome]
                
                total_faltas = int(df_colab_faltas[df_colab_faltas['Tipo'] == 'Falta']['Dias Úteis'].sum())
                total_baixas = int(df_colab_faltas[df_colab_faltas['Tipo'] == 'Baixa']['Dias Úteis'].sum())
                
                dados_colab['Total Faltas (dias)'] = total_faltas
                dados_colab['Total Baixas (dias)'] = total_baixas
                dados_colab['Total Faltas+Baixas'] = total_faltas + total_baixas
            else:
                dados_colab['Total Faltas (dias)'] = 0
                dados_colab['Total Baixas (dias)'] = 0
                dados_colab['Total Faltas+Baixas'] = 0
            
            # Horas Extras
            if not df_horas_extras.empty:
                df_colab_extras = df_horas_extras[df_horas_extras['Nome Completo'] == nome]
                
                dados_colab['Horas Noturnas'] = float(df_colab_extras['Horas Noturnas'].sum())
                dados_colab['Horas Domingos'] = float(df_colab_extras['Horas Domingos'].sum())
                dados_colab['Horas Feriados'] = float(df_colab_extras['Horas Feriados'].sum())
                dados_colab['Horas Extra'] = float(df_colab_extras['Horas Extra'].sum())
                dados_colab['Outros Proveitos'] = float(df_colab_extras['Outros Proveitos'].sum())
                dados_colab['Total Horas Extras'] = (dados_colab['Horas Noturnas'] + 
                                                     dados_colab['Horas Domingos'] + 
                                                     dados_colab['Horas Feriados'] + 
                                                     dados_colab['Horas Extra'])
            else:
                dados_colab['Horas Noturnas'] = 0.0
                dados_colab['Horas Domingos'] = 0.0
                dados_colab['Horas Feriados'] = 0.0
                dados_colab['Horas Extra'] = 0.0
                dados_colab['Outros Proveitos'] = 0.0
                dados_colab['Total Horas Extras'] = 0.0
            
            resultado.append(dados_colab)
        
        df_completo = pd.DataFrame(resultado)
        
        return df_completo
        
    except Exception as e:
        st.error(f"❌ Erro ao carregar dados: {e}")
        return None

def gerar_relatorio_excel(empresa, ano, mes, campos_selecionados, filtros):
    """Gera relatório Excel com campos selecionados"""
    try:
        df_completo = carregar_dados_completos_relatorio(empresa, ano, mes, filtros)
        
        if df_completo is None or df_completo.empty:
            return None
        
        campos_disponiveis = [c for c in campos_selecionados if c in df_completo.columns]
        df_relatorio = df_completo[campos_disponiveis].copy()
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_relatorio.to_excel(writer, sheet_name='Relatório', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Relatório']
            
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output
        
    except Exception as e:
        st.error(f"❌ Erro ao gerar relatório: {e}")
        return None

# ==================== GESTÃO DE TEMPLATES ====================

def salvar_template(nome, campos):
    """Salva um template de relatório"""
    st.session_state.templates_relatorios[nome] = campos
    st.success(f"✅ Template '{nome}' salvo!")

def carregar_template(nome):
    """Carrega um template de relatório"""
    return st.session_state.templates_relatorios.get(nome, [])

def eliminar_template(nome):
    """Elimina um template de relatório"""
    if nome in st.session_state.templates_relatorios:
        del st.session_state.templates_relatorios[nome]
        st.success(f"✅ Template '{nome}' eliminado!")
        return True
    return False

# ==================== FUNÇÕES AUXILIARES PARA FILTROS ====================

def sync_filtros_from_sidebar(prefix):
    """Sincroniza filtros do sidebar para session_state"""
    if f"{prefix}_emp" in st.session_state:
        st.session_state.empresa_selecionada = st.session_state[f"{prefix}_emp"]
    if f"{prefix}_mes" in st.session_state:
        st.session_state.mes_selecionado = st.session_state[f"{prefix}_mes"]
    if f"{prefix}_ano" in st.session_state:
        st.session_state.ano_selecionado = st.session_state[f"{prefix}_ano"]

def criar_filtros_padrao(prefix, incluir_colaborador=True):
    """Cria filtros padrão consistentes"""
    col1, col2, col3 = st.columns(3)
    
    with col1:
        emp_idx = list(EMPRESAS.keys()).index(st.session_state.empresa_selecionada) if st.session_state.empresa_selecionada in EMPRESAS else 0
        emp = st.selectbox("Empresa", list(EMPRESAS.keys()), index=emp_idx, 
                          key=f"{prefix}_emp", on_change=lambda: sync_filtros_from_sidebar(prefix))
        st.session_state.empresa_selecionada = emp
    
    with col2:
        mes = st.selectbox("Mês", list(range(1, 13)),
                          format_func=lambda x: calendar.month_name[x],
                          index=st.session_state.mes_selecionado - 1, 
                          key=f"{prefix}_mes", on_change=lambda: sync_filtros_from_sidebar(prefix))
        st.session_state.mes_selecionado = mes
    
    with col3:
        ano_idx = [2024, 2025, 2026].index(st.session_state.ano_selecionado) if st.session_state.ano_selecionado in [2024, 2025, 2026] else 1
        ano = st.selectbox("Ano", [2024, 2025, 2026], index=ano_idx, 
                          key=f"{prefix}_ano", on_change=lambda: sync_filtros_from_sidebar(prefix))
        st.session_state.ano_selecionado = ano
    
    if incluir_colaborador:
        colabs = carregar_colaboradores_ativos(emp, ano, mes)
        
        if colabs:
            if st.session_state.colaborador_selecionado and st.session_state.colaborador_selecionado in colabs:
                colab_idx = colabs.index(st.session_state.colaborador_selecionado)
            else:
                colab_idx = 0
                st.session_state.colaborador_selecionado = colabs[0]
            
            colab = st.selectbox("Colaborador", colabs, index=colab_idx, key=f"{prefix}_col")
            st.session_state.colaborador_selecionado = colab
            
            return emp, mes, ano, colab
        else:
            return emp, mes, ano, None
    
    return emp, mes, ano

# ==================== INTERFACE ====================

if not check_password():
    st.stop()

st.title("💰 Processamento Salarial v3.5")
st.caption("✨ v3.5: Categoria Profissional + Novos campos pessoais + Filtros Output + CORREÇÃO ROBUSTA ficheiros corrompidos")
st.caption(f"🕐 Reload: {st.session_state.ultimo_reload.strftime('%H:%M:%S')}")

st.markdown("---")

menu = st.sidebar.radio(
    "Menu Principal",
    ["⚙️ Configurações", "💼 Processar Salários", "👥 Visão FTEs/Secção", "📊 Output", "📈 Tabela IRS"],
    index=0
)

# ==================== CONFIGURAÇÕES ====================

if menu == "⚙️ Configurações":
    st.header("⚙️ Configurações do Sistema")
    
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "💶 Feriados", "👥 Colaboradores", "⏰ Horários", 
        "📋 Dados IRS", "🏢 Categoria Profissional", "🔧 Gestão Status"
    ])
    
    with tab1:
        st.subheader("📅 Feriados Municipais")
        feriados_temp = []
        for i in range(3):
            valor_default = st.session_state.feriados_municipais[i] if i < len(st.session_state.feriados_municipais) else None
            feriado = st.date_input(f"Feriado {i+1}", value=valor_default, key=f"fer_{i}")
            if feriado:
                feriados_temp.append(feriado)
        
        if st.button("💾 Atualizar Feriados"):
            st.session_state.feriados_municipais = feriados_temp
            st.success(f"✅ {len(feriados_temp)} feriados")
    
    with tab2:
        st.subheader("👥 Editar Dados dos Colaboradores")
        
        emp, mes, ano, colab = criar_filtros_padrao("cfg", incluir_colaborador=True)
        
        if colab:
            snap = carregar_ultimo_snapshot(emp, colab, ano, mes)
            
            if snap:
                st.markdown("---")
                
                tem_rescisao = snap.get('Data Rescisão', '') != '' and snap.get('Data Rescisão', '') != 'nan'
                
                if tem_rescisao:
                    st.warning(f"⚠️ **Rescisão Registada**: {snap.get('Data Rescisão', 'N/A')}")
                    st.info(f"📋 Motivo: {snap.get('Motivo Rescisão', 'N/A')}")
                    st.markdown("---")
                
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("💰 Salário Bruto", f"{snap['Salário Bruto']:.2f}€")
                col2.metric("🍽️ Subsídio", f"{snap['Subsídio Alimentação Diário']:.2f}€")
                col3.metric("⏰ Horas", f"{snap['Nº Horas/Semana']:.0f}h")
                col4.metric("🔢 Nº Pingo", snap.get('Número Pingo Doce', ''))
                
                col1, col2, col3, col4 = st.columns(4)
                cartao_ref = snap.get('Cartão Refeição', 'Não')
                if cartao_ref == 'Sim':
                    col1.info("💳 Cartão Refeição")
                
                col2.info(f"🏖️ Férias: {snap.get('Sub Férias Tipo', 'Duodécimos')}")
                col3.info(f"🎄 Natal: {snap.get('Sub Natal Tipo', 'Duodécimos')}")
                col4.info(f"🏦 IBAN: {snap.get('IBAN', 'N/A')[:10]}...")
                
                with st.expander("✏️ EDITAR DADOS", expanded=False):
                    with st.form("form_edit"):
                        st.markdown("### 💶 Dados Financeiros")
                        col1, col2 = st.columns(2)
                        with col1:
                            novo_salario = st.number_input("💰 Salário Bruto (€)", min_value=0.0,
                                                          value=float(snap['Salário Bruto']),
                                                          step=10.0, format="%.2f")
                            novo_sub = st.number_input("🍽️ Subsídio Alimentação (€)", min_value=0.0,
                                                      value=float(snap['Subsídio Alimentação Diário']),
                                                      step=0.10, format="%.2f")
                            novo_num = st.text_input("🔢 Número Pingo Doce", value=str(snap.get('Número Pingo Doce', '')))
                        with col2:
                            novo_iban = st.text_input("🏦 IBAN", value=str(snap.get('IBAN', '')),
                                                     help="IBAN bancário do colaborador")
                        
                        st.markdown("### 🏖️ Configurações de Subsídios")
                        col1, col2 = st.columns(2)
                        with col1:
                            sub_ferias_tipo = st.selectbox("Subsídio de Férias", 
                                                          ["Duodécimos", "Total", "Não Pagar"],
                                                          index=["Duodécimos", "Total", "Não Pagar"].index(snap.get('Sub Férias Tipo', 'Duodécimos'))
                                                          if snap.get('Sub Férias Tipo', 'Duodécimos') in ["Duodécimos", "Total", "Não Pagar"] else 0,
                                                          help="Duodécimos = 1/12 por mês | Total = completo num mês | Não Pagar = sem subsídio")
                        with col2:
                            sub_natal_tipo = st.selectbox("Subsídio de Natal", 
                                                         ["Duodécimos", "Total", "Não Pagar"],
                                                         index=["Duodécimos", "Total", "Não Pagar"].index(snap.get('Sub Natal Tipo', 'Duodécimos'))
                                                         if snap.get('Sub Natal Tipo', 'Duodécimos') in ["Duodécimos", "Total", "Não Pagar"] else 0,
                                                         help="Duodécimos = 1/12 por mês | Total = completo num mês | Não Pagar = sem subsídio")
                        
                        st.markdown("### ☑️ Pagamentos")
                        cartao_refeicao = st.checkbox("💳 Pagar em Cartão de Refeição", 
                                                     value=cartao_ref == 'Sim',
                                                     help="Subsídio reconhecido mas pago via cartão (descontado do líquido)")
                        
                        st.markdown("---")
                        st.markdown("### 🚪 Dados de Rescisão")
                        st.info("💡 Registar rescisão aqui mantém o colaborador ATIVO para histórico")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            data_rescisao_valor = snap.get('Data Rescisão', '')
                            if data_rescisao_valor and data_rescisao_valor != '' and data_rescisao_valor != 'nan':
                                try:
                                    data_rescisao_default = datetime.strptime(str(data_rescisao_valor), "%Y-%m-%d").date()
                                except:
                                    data_rescisao_default = None
                            else:
                                data_rescisao_default = None
                            
                            data_rescisao = st.date_input("📅 Data de Rescisão", 
                                                         value=data_rescisao_default,
                                                         help="Deixar vazio se não houver rescisão")
                        
                        with col2:
                            motivo_rescisao_atual = snap.get('Motivo Rescisão', '')
                            motivo_base = motivo_rescisao_atual.split('|')[0].strip() if motivo_rescisao_atual else ''
                            
                            if motivo_base and motivo_base in MOTIVOS_RESCISAO:
                                motivo_idx = MOTIVOS_RESCISAO.index(motivo_base)
                            else:
                                motivo_idx = 0
                            
                            motivo_rescisao = st.selectbox("📋 Motivo da Rescisão", 
                                                          MOTIVOS_RESCISAO,
                                                          index=motivo_idx)
                        
                        obs_rescisao = st.text_area("📝 Observações da Rescisão", 
                                                   value="",
                                                   help="Detalhes adicionais sobre a rescisão")
                        
                        submit = st.form_submit_button("💾 GUARDAR TUDO", use_container_width=True, type="primary")
                        
                        if submit:
                            df_base = carregar_dados_base(emp)
                            excel_file = download_excel(emp)
                            wb = load_workbook(excel_file, data_only=False)
                            
                            mask = df_base['Nome Completo'] == colab
                            df_base.loc[mask, 'Salário Bruto'] = novo_salario
                            df_base.loc[mask, 'Subsídio Alimentação Diário'] = novo_sub
                            df_base.loc[mask, 'Número Pingo Doce'] = novo_num
                            df_base.loc[mask, 'IBAN'] = novo_iban
                            df_base.loc[mask, 'Cartão Refeição'] = 'Sim' if cartao_refeicao else 'Não'
                            df_base.loc[mask, 'Sub Férias Tipo'] = sub_ferias_tipo
                            df_base.loc[mask, 'Sub Natal Tipo'] = sub_natal_tipo
                            
                            if data_rescisao:
                                df_base.loc[mask, 'Data Rescisão'] = data_rescisao.strftime("%Y-%m-%d")
                                obs_completa = motivo_rescisao
                                if obs_rescisao:
                                    obs_completa += f" | Obs: {obs_rescisao}"
                                df_base.loc[mask, 'Motivo Rescisão'] = obs_completa
                            
                            if not validar_workbook(wb):
                                st.error("❌ Validação falhou")
                                st.stop()
                            
                            if "Colaboradores" in wb.sheetnames:
                                idx = wb.sheetnames.index("Colaboradores")
                                del wb["Colaboradores"]
                                ws = wb.create_sheet("Colaboradores", idx)
                            else:
                                ws = wb.create_sheet("Colaboradores", 0)
                            
                            for r in dataframe_to_rows(df_base, index=False, header=True):
                                ws.append(r)
                            
                            if upload_excel_seguro(emp, wb):
                                st.success("✅ Todos os dados atualizados!")
                                if data_rescisao:
                                    st.info(f"🚪 Rescisão registada: {data_rescisao.strftime('%d/%m/%Y')}")
                                st.balloons()
                                time.sleep(2)
                                st.rerun()
        else:
            st.warning("⚠️ Nenhum colaborador ativo")
    
    with tab3:
        st.subheader("⏰ Mudanças de Horário")
        
        emp, mes, ano, colab = criar_filtros_padrao("hor", incluir_colaborador=True)
        
        if colab:
            snap_hor = carregar_ultimo_snapshot(emp, colab, ano, mes)
            
            if snap_hor:
                st.markdown("---")
                
                col1, col2, col3 = st.columns(3)
                horas_atuais = float(snap_hor['Nº Horas/Semana'])
                salario_atual = float(snap_hor['Salário Bruto'])
                venc_hora_atual = float(snap_hor['Vencimento Hora'])
                
                col1.metric("⏰ Horas Atuais", f"{horas_atuais:.0f}h/semana")
                col2.metric("💰 Salário Bruto", f"{salario_atual:.2f}€")
                col3.metric("💵 Vencimento/Hora", f"{venc_hora_atual:.2f}€")
                
                st.markdown("---")
                
                with st.form("form_horario"):
                    st.markdown("### Novo Horário")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        novas_horas = st.selectbox(
                            "⏰ Novas Horas Semanais",
                            options=HORAS_PERMITIDAS,
                            index=HORAS_PERMITIDAS.index(int(horas_atuais)) if int(horas_atuais) in HORAS_PERMITIDAS else 2
                        )
                    
                    with col2:
                        novo_venc_hora = calcular_vencimento_hora(salario_atual, novas_horas)
                        
                        st.metric("💵 Novo Vencimento/Hora", f"{novo_venc_hora:.2f}€",
                                 delta=f"{novo_venc_hora - venc_hora_atual:.2f}€")
                        st.caption("(Salário mantém-se)")
                    
                    submit_hor = st.form_submit_button("💾 CONFIRMAR", use_container_width=True, type="primary")
                    
                    if submit_hor:
                        if novas_horas == horas_atuais:
                            st.warning("⚠️ As horas não foram alteradas!")
                        else:
                            df_base = carregar_dados_base(emp)
                            excel_file = download_excel(emp)
                            wb = load_workbook(excel_file, data_only=False)
                            
                            if not validar_workbook(wb):
                                st.error("❌ Validação falhou")
                                st.stop()
                            
                            mask = df_base['Nome Completo'] == colab
                            df_base.loc[mask, 'Nº Horas/Semana'] = novas_horas
                            
                            if "Colaboradores" in wb.sheetnames:
                                idx = wb.sheetnames.index("Colaboradores")
                                del wb["Colaboradores"]
                                ws = wb.create_sheet("Colaboradores", idx)
                            else:
                                ws = wb.create_sheet("Colaboradores", 0)
                            
                            for r in dataframe_to_rows(df_base, index=False, header=True):
                                ws.append(r)
                            
                            if upload_excel_seguro(emp, wb):
                                st.success("✅ Horário atualizado!")
                                st.balloons()
                                time.sleep(2)
                                st.rerun()
        else:
            st.warning("⚠️ Nenhum colaborador ativo")
    
    with tab4:
        st.subheader("📋 Dados IRS")
        
        emp, mes, ano, colab = criar_filtros_padrao("irs", incluir_colaborador=True)
        
        if colab:
            snap_irs = carregar_ultimo_snapshot(emp, colab, ano, mes)
            
            if snap_irs:
                with st.form("form_irs"):
                    st.markdown(f"### {colab}")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        estado_civil = st.selectbox("Estado Civil", ESTADOS_CIVIS,
                                                   index=ESTADOS_CIVIS.index(snap_irs.get('Estado Civil', 'Solteiro'))
                                                   if snap_irs.get('Estado Civil') in ESTADOS_CIVIS else 0)
                        num_titulares = st.number_input("Nº Titulares", min_value=1, max_value=2,
                                                       value=int(snap_irs.get('Nº Titulares', 2)))
                        num_dependentes = st.number_input("Nº Dependentes", min_value=0,
                                                         value=int(snap_irs.get('Nº Dependentes', 0)))
                    
                    with col2:
                        tem_deficiencia = st.selectbox("Deficiência", ["Não", "Sim"],
                                                      index=0 if snap_irs.get('Deficiência', 'Não') == 'Não' else 1)
                        irs_modo = st.selectbox("Modo Cálculo IRS", ["Tabela", "Fixa"],
                                               index=0 if snap_irs.get('IRS Modo Calculo', 'Tabela') == 'Tabela' else 1)
                        irs_percentagem = st.number_input("IRS % Fixa", min_value=0.0, max_value=100.0,
                                                         value=float(snap_irs.get('IRS Percentagem Fixa', 0)),
                                                         step=0.1, format="%.1f")
                    
                    submit_irs = st.form_submit_button("💾 GUARDAR", use_container_width=True, type="primary")
                    
                    if submit_irs:
                        df_base = carregar_dados_base(emp)
                        excel_file = download_excel(emp)
                        wb = load_workbook(excel_file, data_only=False)
                        
                        if not validar_workbook(wb):
                            st.error("❌ Validação falhou")
                            st.stop()
                        
                        mask = df_base['Nome Completo'] == colab
                        df_base.loc[mask, 'Estado Civil'] = estado_civil
                        df_base.loc[mask, 'Nº Titulares'] = num_titulares
                        df_base.loc[mask, 'Nº Dependentes'] = num_dependentes
                        df_base.loc[mask, 'Pessoa com Deficiência'] = tem_deficiencia
                        df_base.loc[mask, 'Tipo IRS'] = irs_modo
                        df_base.loc[mask, '% IRS Fixa'] = irs_percentagem
                        
                        if "Colaboradores" in wb.sheetnames:
                            idx = wb.sheetnames.index("Colaboradores")
                            del wb["Colaboradores"]
                            ws = wb.create_sheet("Colaboradores", idx)
                        else:
                            ws = wb.create_sheet("Colaboradores", 0)
                        
                        for r in dataframe_to_rows(df_base, index=False, header=True):
                            ws.append(r)
                        
                        if upload_excel_seguro(emp, wb):
                            st.success("✅ Dados IRS atualizados!")
                            st.balloons()
                            time.sleep(2)
                            st.rerun()
        else:
            st.warning("⚠️ Nenhum colaborador ativo")
    
    with tab5:
        st.subheader("🏢 Categoria Profissional")
        st.caption("✨ v3.5: Nova funcionalidade")
        
        emp, mes, ano, colab = criar_filtros_padrao("cat", incluir_colaborador=True)
        
        if colab:
            snap_cat = carregar_ultimo_snapshot(emp, colab, ano, mes)
            
            if snap_cat:
                st.markdown("---")
                
                cat_atual = snap_cat.get('Categoria Profissional', '')
                
                if cat_atual:
                    st.info(f"📋 Categoria Atual: **{cat_atual}**")
                else:
                    st.warning("⚠️ Sem categoria definida")
                
                st.markdown("---")
                
                with st.form("form_categoria"):
                    st.markdown(f"### Definir Categoria Profissional - {colab}")
                    
                    # Obter categorias disponíveis para a empresa
                    categorias_empresa = CATEGORIAS_PROFISSIONAIS.get(emp, [])
                    
                    if not categorias_empresa:
                        st.warning(f"⚠️ Sem categorias definidas para {emp}")
                        st.stop()
                    
                    # Encontrar índice da categoria atual
                    if cat_atual and cat_atual in categorias_empresa:
                        cat_idx = categorias_empresa.index(cat_atual)
                    else:
                        cat_idx = 0
                    
                    nova_categoria = st.selectbox(
                        "🏢 Categoria Profissional",
                        options=categorias_empresa,
                        index=cat_idx,
                        help=f"Categorias disponíveis para {emp}"
                    )
                    
                    obs_categoria = st.text_area(
                        "📝 Observações",
                        help="Observações sobre a categoria ou mudança (opcional)"
                    )
                    
                    submit_cat = st.form_submit_button("💾 GUARDAR", use_container_width=True, type="primary")
                    
                    if submit_cat:
                        df_base = carregar_dados_base(emp)
                        excel_file = download_excel(emp)
                        wb = load_workbook(excel_file, data_only=False)
                        
                        if not validar_workbook(wb):
                            st.error("❌ Validação falhou")
                            st.stop()
                        
                        mask = df_base['Nome Completo'] == colab
                        df_base.loc[mask, 'Categoria Profissional'] = nova_categoria
                        
                        if "Colaboradores" in wb.sheetnames:
                            idx = wb.sheetnames.index("Colaboradores")
                            del wb["Colaboradores"]
                            ws = wb.create_sheet("Colaboradores", idx)
                        else:
                            ws = wb.create_sheet("Colaboradores", 0)
                        
                        for r in dataframe_to_rows(df_base, index=False, header=True):
                            ws.append(r)
                        
                        if upload_excel_seguro(emp, wb):
                            st.success(f"✅ Categoria atualizada: **{nova_categoria}**")
                            if obs_categoria:
                                st.info(f"📝 Observação: {obs_categoria}")
                            st.balloons()
                            time.sleep(2)
                            st.rerun()
        else:
            st.warning("⚠️ Nenhum colaborador ativo")
    
    with tab6:
        st.subheader("🔧 Gestão de Status dos Colaboradores")
        
        col1, col2 = st.columns(2)
        with col1:
            emp_idx_status = list(EMPRESAS.keys()).index(st.session_state.empresa_selecionada) if st.session_state.empresa_selecionada in EMPRESAS else 0
            emp_status = st.selectbox("Empresa", list(EMPRESAS.keys()), index=emp_idx_status, key="emp_status")
            st.session_state.empresa_selecionada = emp_status
        with col2:
            mostrar = st.radio("Mostrar", ["Ativos", "Inativos", "Todos"], horizontal=True)
        
        df_base = carregar_dados_base(emp_status)
        
        if not df_base.empty:
            if mostrar == "Ativos":
                df_filtrado = df_base[df_base['Status'] == 'Ativo']
            elif mostrar == "Inativos":
                df_filtrado = df_base[df_base['Status'] != 'Ativo']
            else:
                df_filtrado = df_base
            
            st.markdown(f"**Total: {len(df_filtrado)} colaboradores**")
            st.markdown("---")
            
            if not df_filtrado.empty:
                for _, row in df_filtrado.iterrows():
                    nome = row['Nome Completo']
                    status_atual = row.get('Status', 'Ativo')
                    
                    col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
                    
                    with col1:
                        st.write(f"**{nome}**")
                        info_text = f"Secção: {row.get('Secção', 'N/A')} | Salário: {row.get('Salário Bruto', 0):.2f}€"
                        
                        if row.get('Data Rescisão', '') and row.get('Data Rescisão', '') != 'nan':
                            info_text += f" | 🚪 Rescisão: {row.get('Data Rescisão', 'N/A')}"
                        
                        st.caption(info_text)
                    
                    with col2:
                        if status_atual == 'Ativo':
                            st.success("✅ Ativo")
                        else:
                            st.error("❌ Inativo")
                    
                    with col3:
                        if status_atual == 'Ativo':
                            if st.button("❌ Desativar", key=f"desativar_{nome}"):
                                if atualizar_status_colaborador(emp_status, nome, 'Inativo'):
                                    st.rerun()
                    
                    with col4:
                        if status_atual != 'Ativo':
                            if st.button("✅ Ativar", key=f"ativar_{nome}"):
                                if atualizar_status_colaborador(emp_status, nome, 'Ativo'):
                                    st.rerun()
                    
                    st.markdown("---")
            else:
                st.info("ℹ️ Nenhum colaborador encontrado")

# ==================== PROCESSAR SALÁRIOS ====================

elif menu == "💼 Processar Salários":
    st.header("💼 Processamento Mensal")
    
    emp, mes, ano, colab = criar_filtros_padrao("proc", incluir_colaborador=True)
    
    if not colab:
        st.warning("⚠️ Nenhum colaborador ativo")
        st.stop()
    
    snap_proc = carregar_ultimo_snapshot(emp, colab, ano, mes)
    
    if not snap_proc:
        st.error("❌ Erro ao carregar")
        st.stop()
    
    salario_bruto = float(snap_proc['Salário Bruto'])
    horas_semana = float(snap_proc['Nº Horas/Semana'])
    subsidio_alim = float(snap_proc['Subsídio Alimentação Diário'])
    vencimento_hora = float(snap_proc['Vencimento Hora'])
    
    feriados = FERIADOS_NACIONAIS_2025 + st.session_state.feriados_municipais
    dias_uteis_mes = calcular_dias_uteis(ano, mes, feriados)
    
    st.markdown("---")
    
    with st.expander("📋 DADOS BASE", expanded=True):
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("💰 Salário Bruto", f"{salario_bruto:.2f}€")
        col2.metric("⏰ Horas/Semana", f"{horas_semana:.0f}h")
        col3.metric("💵 Vencimento/Hora", f"{vencimento_hora:.2f}€")
        col4.metric("🍽️ Sub. Alimentação", f"{subsidio_alim:.2f}€/dia")
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("📅 Dias Úteis Mês", dias_uteis_mes)
        col2.metric("👤 Estado Civil", snap_proc.get('Estado Civil', 'N/A'))
        col3.metric("👶 Dependentes", snap_proc.get('Nº Dependentes', 0))
        col4.metric("📊 Modo IRS", snap_proc.get('IRS Modo Calculo', 'Tabela'))
        
        col1, col2, col3 = st.columns(3)
        if snap_proc.get('Cartão Refeição', 'Não') == 'Sim':
            col1.info("💳 Cartão Refeição")
        col2.info(f"🏖️ {snap_proc.get('Sub Férias Tipo', 'Duodécimos')}")
        col3.info(f"🎄 {snap_proc.get('Sub Natal Tipo', 'Duodécimos')}")
    
    st.markdown("---")
    
    # FALTAS E BAIXAS
    st.subheader("🏖️ Faltas e Baixas Médicas")
    
    tab_faltas, tab_baixas, tab_historico = st.tabs(["➕ Nova Falta", "🏥 Nova Baixa", "📜 Histórico"])
    
    with tab_faltas:
        with st.form("form_nova_falta"):
            st.markdown("### Registar Falta")
            
            col1, col2 = st.columns(2)
            with col1:
                data_inicio_falta = st.date_input("📅 Data Início", value=date.today(), key="falta_inicio")
            with col2:
                data_fim_falta = st.date_input("📅 Data Fim", value=date.today(), key="falta_fim")
            
            obs_falta = st.text_area("📝 Observações", key="obs_falta")
            
            submit_falta = st.form_submit_button("💾 REGISTAR FALTA", use_container_width=True, type="primary")
            
            if submit_falta:
                if data_inicio_falta > data_fim_falta:
                    st.error("❌ Data de início deve ser anterior à data de fim!")
                else:
                    with st.spinner("A registar..."):
                        if gravar_falta_baixa(emp, ano, mes, colab, 
                                            "Falta", data_inicio_falta, data_fim_falta, obs_falta):
                            st.success("✅ Falta registada!")
                            time.sleep(1)
                            st.rerun()
    
    with tab_baixas:
        with st.form("form_nova_baixa"):
            st.markdown("### Registar Baixa Médica")
            
            col1, col2 = st.columns(2)
            with col1:
                data_inicio_baixa = st.date_input("📅 Data Início", value=date.today(), key="baixa_inicio")
            with col2:
                data_fim_baixa = st.date_input("📅 Data Fim", value=date.today(), key="baixa_fim")
            
            obs_baixa = st.text_area("📝 Observações", key="obs_baixa")
            
            ficheiro_baixa = st.file_uploader("📎 Anexar Documento", 
                                             type=['pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx'],
                                             key="file_baixa")
            
            submit_baixa = st.form_submit_button("💾 REGISTAR BAIXA", use_container_width=True, type="primary")
            
            if submit_baixa:
                if data_inicio_baixa > data_fim_baixa:
                    st.error("❌ Data de início deve ser anterior à data de fim!")
                else:
                    with st.spinner("A processar..."):
                        ficheiro_path = None
                        
                        if ficheiro_baixa:
                            st.info("📤 A fazer upload...")
                            ficheiro_path = upload_ficheiro_baixa(emp, ano, mes, 
                                                                 colab, ficheiro_baixa)
                            if ficheiro_path:
                                st.success(f"✅ Documento guardado")
                        
                        if gravar_falta_baixa(emp, ano, mes, colab,
                                            "Baixa", data_inicio_baixa, data_fim_baixa, 
                                            obs_baixa, ficheiro_path):
                            st.success("✅ Baixa registada!")
                            time.sleep(2)
                            st.rerun()
    
    with tab_historico:
        st.markdown("### 📜 Histórico do Mês")
        
        df_historico = carregar_faltas_baixas(emp, ano, mes, colab)
        
        if not df_historico.empty:
            for idx, row in df_historico.iterrows():
                col1, col2, col3, col4, col5, col6, col7 = st.columns([2, 1, 1, 1, 1, 2, 1])
                
                with col1:
                    icone = "🏖️" if row['Tipo'] == 'Falta' else "🏥"
                    st.write(f"{icone} **{row['Tipo']}**")
                with col2:
                    st.caption(f"📅 {row['Data Início']}")
                with col3:
                    st.caption(f"→ {row['Data Fim']}")
                with col4:
                    st.caption(f"📊 {row['Dias Úteis']} úteis")
                with col5:
                    st.caption(f"📅 {row['Dias Totais']} totais")
                with col6:
                    if row['Observações']:
                        st.caption(f"📝 {row['Observações']}")
                with col7:
                    if st.button("🗑️", key=f"del_fb_{idx}", help="Eliminar registo"):
                        if eliminar_registo_falta_baixa(emp, ano, mes, idx):
                            time.sleep(1)
                            st.rerun()
                
                st.markdown("---")
            
            total_faltas_uteis = df_historico[df_historico['Tipo'] == 'Falta']['Dias Úteis'].sum()
            total_baixas_uteis = df_historico[df_historico['Tipo'] == 'Baixa']['Dias Úteis'].sum()
            
            col1, col2, col3 = st.columns(3)
            col1.metric("🏖️ Total Faltas", int(total_faltas_uteis))
            col2.metric("🏥 Total Baixas", int(total_baixas_uteis))
            col3.metric("📊 Total Geral", int(total_faltas_uteis + total_baixas_uteis))
        else:
            st.info("ℹ️ Sem registos")
    
    st.markdown("---")
    
    # CALCULAR DIAS ÚTEIS TRABALHADOS
    df_historico = carregar_faltas_baixas(emp, ano, mes, colab)
    
    if not df_historico.empty:
        total_faltas_uteis = int(df_historico[df_historico['Tipo'] == 'Falta']['Dias Úteis'].sum())
        total_baixas_uteis = int(df_historico[df_historico['Tipo'] == 'Baixa']['Dias Úteis'].sum())
    else:
        total_faltas_uteis = 0
        total_baixas_uteis = 0
    
    dias_uteis_trab = max(dias_uteis_mes - total_faltas_uteis - total_baixas_uteis, 0)
    
    st.info(f"📊 Dias úteis trabalhados: {dias_uteis_trab} de {dias_uteis_mes}")
    
    st.markdown("---")
    
    # HORAS EXTRAS
    st.subheader("⏰ Horas Extras e Outros Proveitos")
    
    tab_registar, tab_historico_extras = st.tabs(["➕ Registar", "📜 Histórico"])
    
    with tab_registar:
        with st.form("form_horas_extras"):
            st.markdown("### Registar Horas Extras")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                h_not = st.number_input("🌙 Noturnas", min_value=0.0, value=0.0, step=0.5, key="reg_h_not")
            with col2:
                h_dom = st.number_input("📅 Domingos", min_value=0.0, value=0.0, step=0.5, key="reg_h_dom")
            with col3:
                h_fer = st.number_input("🎉 Feriados", min_value=0.0, value=0.0, step=0.5, key="reg_h_fer")
            with col4:
                h_ext = st.number_input("⚡ Extra", min_value=0.0, value=0.0, step=0.5, key="reg_h_ext")
            
            outros_prov = st.number_input("💰 Outros Proveitos (€)", min_value=0.0, value=0.0, key="reg_outros")
            obs_extras = st.text_area("📝 Observações", key="obs_extras")
            
            submit_extras = st.form_submit_button("💾 REGISTAR", use_container_width=True, type="primary")
            
            if submit_extras:
                if h_not == 0 and h_dom == 0 and h_fer == 0 and h_ext == 0 and outros_prov == 0:
                    st.warning("⚠️ Nenhum valor preenchido!")
                else:
                    with st.spinner("A registar..."):
                        if gravar_horas_extras(emp, ano, mes, colab,
                                              h_not, h_dom, h_fer, h_ext, outros_prov, obs_extras):
                            st.success("✅ Extras registados!")
                            time.sleep(1)
                            st.rerun()
    
    with tab_historico_extras:
        st.markdown("### 📜 Histórico de Extras")
        
        df_extras = carregar_horas_extras(emp, ano, mes, colab)
        
        if not df_extras.empty:
            for idx, row in df_extras.iterrows():
                col1, col2, col3, col4, col5, col6, col7, col8 = st.columns([1, 1, 1, 1, 1, 1, 2, 1])
                
                with col1:
                    st.caption(f"🌙 {row['Horas Noturnas']:.1f}h")
                with col2:
                    st.caption(f"📅 {row['Horas Domingos']:.1f}h")
                with col3:
                    st.caption(f"🎉 {row['Horas Feriados']:.1f}h")
                with col4:
                    st.caption(f"⚡ {row['Horas Extra']:.1f}h")
                with col5:
                    st.caption(f"💰 {row['Outros Proveitos']:.2f}€")
                with col6:
                    st.caption(f"📅 {row['Timestamp'][:10]}")
                with col7:
                    if row['Observações']:
                        st.caption(f"📝 {row['Observações']}")
                with col8:
                    if st.button("🗑️", key=f"del_ext_{idx}", help="Eliminar registo"):
                        if eliminar_registo_horas_extras(emp, ano, mes, idx):
                            time.sleep(1)
                            st.rerun()
                
                st.markdown("---")
            
            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("🌙 Noturnas", f"{df_extras['Horas Noturnas'].sum():.1f}h")
            col2.metric("📅 Domingos", f"{df_extras['Horas Domingos'].sum():.1f}h")
            col3.metric("🎉 Feriados", f"{df_extras['Horas Feriados'].sum():.1f}h")
            col4.metric("⚡ Extra", f"{df_extras['Horas Extra'].sum():.1f}h")
            col5.metric("💰 Proveitos", f"{df_extras['Outros Proveitos'].sum():.2f}€")
        else:
            st.info("ℹ️ Sem registos")
    
    st.markdown("---")
    
    # CARREGAR VALORES DO HISTÓRICO
    df_extras_total = carregar_horas_extras(emp, ano, mes, colab)
    
    if not df_extras_total.empty:
        h_not_total = float(df_extras_total['Horas Noturnas'].sum())
        h_dom_total = float(df_extras_total['Horas Domingos'].sum())
        h_fer_total = float(df_extras_total['Horas Feriados'].sum())
        h_ext_total = float(df_extras_total['Horas Extra'].sum())
        outros_prov_total = float(df_extras_total['Outros Proveitos'].sum())
    else:
        h_not_total = 0.0
        h_dom_total = 0.0
        h_fer_total = 0.0
        h_ext_total = 0.0
        outros_prov_total = 0.0
    
    # CONFIGURAÇÕES
    cartao_ref_ativo = snap_proc.get('Cartão Refeição', 'Não') == 'Sim'
    sub_ferias_tipo = snap_proc.get('Sub Férias Tipo', 'Duodécimos')
    sub_natal_tipo = snap_proc.get('Sub Natal Tipo', 'Duodécimos')
    
    st.markdown("---")
    
    dados_calc = {
        'salario_bruto': salario_bruto,
        'horas_semana': horas_semana,
        'subsidio_alimentacao': subsidio_alim,
        'dias_faltas': total_faltas_uteis,
        'dias_baixas': total_baixas_uteis,
        'dias_uteis_trabalhados': dias_uteis_trab,
        'horas_noturnas': h_not_total,
        'horas_domingos': h_dom_total,
        'horas_feriados': h_fer_total,
        'horas_extra': h_ext_total,
        'sub_ferias_tipo': sub_ferias_tipo,
        'sub_natal_tipo': sub_natal_tipo,
        'cartao_refeicao': cartao_ref_ativo,
        'outros_proveitos': outros_prov_total,
        'estado_civil': snap_proc.get('Estado Civil', 'Solteiro'),
        'num_dependentes': snap_proc.get('Nº Dependentes', 0),
        'tem_deficiencia': snap_proc.get('Deficiência', 'Não') == 'Sim',
        'irs_modo': snap_proc.get('IRS Modo Calculo', 'Tabela'),
        'irs_percentagem_fixa': snap_proc.get('IRS Percentagem Fixa', 0)
    }
    
    resultado = processar_calculo_salario(dados_calc)
    
    st.subheader("💵 Preview do Processamento")
    
    # Tabelas de resumo
    dados_remuneracoes = {
        "Descrição": [
            "Vencimento Ajustado",
            "Sub. Alimentação",
            "Trabalho Noturno (+25%)",
            "Domingos",
            "Feriados (×2)",
            "Sub. Férias",
            "Sub. Natal",
            "Horas Extra"
        ],
        "Valor (€)": [
            f"{resultado['vencimento_ajustado']:.2f}",
            f"{resultado['sub_alimentacao']:.2f}",
            f"{resultado['trabalho_noturno']:.2f}",
            f"{resultado['domingos']:.2f}",
            f"{resultado['feriados']:.2f}",
            f"{resultado['sub_ferias']:.2f}",
            f"{resultado['sub_natal']:.2f}",
            f"{resultado['banco_horas_valor']:.2f}"
        ]
    }
    
    if outros_prov_total > 0:
        dados_remuneracoes["Descrição"].append("Outros Proveitos")
        dados_remuneracoes["Valor (€)"].append(f"{resultado['outros_proveitos']:.2f}")
    
    dados_descontos = {
        "Descrição": [
            "Base Seg. Social",
            "Segurança Social (11%)",
            "Base IRS",
            "IRS"
        ],
        "Valor (€)": [
            f"{resultado['base_ss']:.2f}",
            f"{resultado['seg_social']:.2f}",
            f"{resultado['base_irs']:.2f}",
            f"{resultado['irs']:.2f}"
        ]
    }
    
    if resultado.get('cartao_refeicao', False) and resultado['desconto_especie'] > 0:
        dados_descontos["Descrição"].append("💳 Cartão Refeição")
        dados_descontos["Valor (€)"].append(f"{resultado['desconto_especie']:.2f}")
    
    dados_resumo = {
        "Descrição": [
            "Dias Úteis do Mês",
            "Faltas (dias úteis)",
            "Baixas (dias úteis)",
            "Dias Úteis Trabalhados",
            "Dias Pagos (sobre 30)"
        ],
        "Valor": [
            str(dias_uteis_mes),
            str(total_faltas_uteis),
            str(total_baixas_uteis),
            str(dias_uteis_trab),
            str(30 - total_faltas_uteis - total_baixas_uteis)
        ]
    }
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("### 💰 Remunerações")
        df_rem = pd.DataFrame(dados_remuneracoes)
        st.dataframe(df_rem, hide_index=True, use_container_width=True)
        st.markdown(f"**TOTAL: {resultado['total_remuneracoes']:.2f}€**")
    
    with col2:
        st.markdown("### 📉 Descontos")
        df_desc = pd.DataFrame(dados_descontos)
        st.dataframe(df_desc, hide_index=True, use_container_width=True)
        st.markdown(f"**TOTAL: {resultado['total_descontos']:.2f}€**")
    
    with col3:
        st.markdown("### 💵 Resumo")
        df_resumo = pd.DataFrame(dados_resumo)
        st.dataframe(df_resumo, hide_index=True, use_container_width=True)
        st.markdown(f"### **💰 {resultado['liquido']:.2f}€**")

# ==================== VISÃO FTEs ====================

elif menu == "👥 Visão FTEs/Secção":
    st.header("👥 Visão de FTEs por Secção")
    
    emp, mes, ano = criar_filtros_padrao("ftes", incluir_colaborador=False)
    
    st.markdown("---")
    
    stats = calcular_ftes_e_estatisticas(emp, ano, mes)
    
    if stats:
        col1, col2 = st.columns(2)
        with col1:
            st.metric("👥 Total de Colaboradores Ativos", stats['total_colaboradores'])
        with col2:
            st.metric("📊 Total de FTEs (Full-Time Equivalents)", stats['total_ftes'])
        
        st.caption(f"💡 FTEs = Total de horas semanais ÷ 40 = {stats['total_horas']:.0f}h ÷ 40 = {stats['total_ftes']}")
        
        st.markdown("---")
        
        st.subheader("📋 Detalhes por Secção")
        
        df_display = stats['df_stats'].copy()
        
        st.dataframe(
            df_display,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Secção": st.column_config.TextColumn("Secção", width="medium"),
                "Nº Colaboradores": st.column_config.NumberColumn("Nº Colab.", width="small"),
                "16h": st.column_config.NumberColumn("16h", width="small"),
                "20h": st.column_config.NumberColumn("20h", width="small"),
                "40h": st.column_config.NumberColumn("40h", width="small"),
                "Total Horas/Semana": st.column_config.NumberColumn("Total Horas", width="medium", format="%.0f"),
                "FTEs": st.column_config.NumberColumn("FTEs", width="small", format="%.2f")
            }
        )
        
        st.markdown("---")
        
        st.subheader("📊 Visualizações")
        
        seccoes_disponiveis = ['Todas'] + sorted(df_display['Secção'].unique().tolist())
        seccao_filtro = st.selectbox("🔍 Filtrar por Secção", seccoes_disponiveis, key="filtro_seccao_viz")
        
        if seccao_filtro != 'Todas':
            df_viz = df_display[df_display['Secção'] == seccao_filtro].copy()
        else:
            df_viz = df_display.copy()
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### Colaboradores por Secção")
            chart_data = df_viz[['Secção', 'Nº Colaboradores']].set_index('Secção')
            st.bar_chart(chart_data)
        
        with col2:
            st.markdown("#### FTEs por Secção")
            chart_data = df_viz[['Secção', 'FTEs']].set_index('Secção')
            st.bar_chart(chart_data)
        
    else:
        st.warning("⚠️ Sem dados disponíveis")

# ==================== OUTPUT ====================

elif menu == "📊 Output":
    st.header("📊 Exportação de Relatórios")
    
    emp, mes, ano = criar_filtros_padrao("output", incluir_colaborador=False)
    
    st.markdown("---")
    
    df_completo = carregar_dados_completos_relatorio(emp, ano, mes, {})
    
    if df_completo is None or df_completo.empty:
        st.warning("⚠️ Sem dados disponíveis")
        st.stop()
    
    st.subheader("🎯 Configurar Relatório")
    
    # GESTÃO DE TEMPLATES
    with st.expander("📁 GESTÃO DE TEMPLATES", expanded=False):
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown("### 💾 Templates Salvos")
            
            if st.session_state.templates_relatorios:
                for nome_template in st.session_state.templates_relatorios.keys():
                    col_a, col_b, col_c = st.columns([3, 1, 1])
                    
                    with col_a:
                        st.write(f"📋 **{nome_template}**")
                        st.caption(f"{len(st.session_state.templates_relatorios[nome_template])} campos")
                    
                    with col_b:
                        if st.button("📥 Carregar", key=f"load_{nome_template}"):
                            st.session_state.campos_template_carregado = nome_template
                            st.rerun()
                    
                    with col_c:
                        if st.button("🗑️", key=f"del_tmpl_{nome_template}"):
                            eliminar_template(nome_template)
                            time.sleep(1)
                            st.rerun()
                    
                    st.markdown("---")
            else:
                st.info("ℹ️ Nenhum template salvo")
        
        with col2:
            st.markdown("### ➕ Novo Template")
            
            nome_novo_template = st.text_input("Nome do Template", 
                                              placeholder="ex: Contabilidade",
                                              key="nome_novo_template")
            
            if st.button("💾 Salvar Template Atual", use_container_width=True):
                if nome_novo_template:
                    if st.session_state.campos_selecionados_output:
                        salvar_template(nome_novo_template, st.session_state.campos_selecionados_output)
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.warning("⚠️ Selecione campos primeiro!")
                else:
                    st.warning("⚠️ Digite um nome para o template!")
    
    st.markdown("---")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### 📋 Selecionar Campos por Categoria")
        
        # v3.5: Botão Limpar Campos
        if st.button("🗑️ Limpar Todos os Campos", use_container_width=True, type="secondary"):
            st.session_state.campos_selecionados_output = []
            st.success("✅ Campos limpos!")
            time.sleep(1)
            st.rerun()
        
        st.markdown("---")
        
        # Verificar se há template carregado
        campos_pre_selecionados = []
        if 'campos_template_carregado' in st.session_state:
            nome_template_carregado = st.session_state.campos_template_carregado
            campos_pre_selecionados = carregar_template(nome_template_carregado)
            st.info(f"📥 Template carregado: **{nome_template_carregado}**")
            del st.session_state.campos_template_carregado
        
        campos_selecionados = []
        
        # Iterar por cada categoria
        for categoria, campos_categoria in CATEGORIAS_CAMPOS.items():
            with st.expander(categoria, expanded=False):
                # Filtrar campos disponíveis na categoria
                campos_disponiveis = [c for c in campos_categoria if c in df_completo.columns]
                
                if campos_disponiveis:
                    # Determinar campos default
                    if campos_pre_selecionados:
                        campos_default = [c for c in campos_disponiveis if c in campos_pre_selecionados]
                    elif st.session_state.campos_selecionados_output:
                        campos_default = [c for c in campos_disponiveis if c in st.session_state.campos_selecionados_output]
                    else:
                        # Pré-selecionar campos básicos
                        campos_basicos = ['Nome Completo', 'Status', 'Secção', 'Salário Bruto']
                        campos_default = [c for c in campos_disponiveis if c in campos_basicos]
                    
                    # Multiselect para cada categoria
                    campos_cat = st.multiselect(
                        f"Campos de {categoria}",
                        options=campos_disponiveis,
                        default=campos_default,
                        key=f"campos_{categoria}",
                        label_visibility="collapsed"
                    )
                    
                    campos_selecionados.extend(campos_cat)
        
        # Salvar no session_state
        st.session_state.campos_selecionados_output = campos_selecionados
    
    with col2:
        st.markdown("### 🔍 Filtros")
        
        # v3.5: Status predefinido "Ativo"
        status_filtro = st.selectbox(
            "Status",
            ["Ativo", "Inativo", "Todos"],
            index=0,  # "Ativo" por defeito
            help="Filtrar por status do colaborador"
        )
        
        seccoes = ['Todas'] + sorted(df_completo['Secção'].dropna().unique().tolist())
        seccao_filtro = st.selectbox(
            "Secção",
            seccoes,
            help="Filtrar por secção"
        )
        
        # v3.5: Filtro por Colaborador (múltipla seleção)
        st.markdown("**Colaboradores**")
        todos_colaboradores = sorted(df_completo['Nome Completo'].unique().tolist())
        
        colaboradores_selecionados = st.multiselect(
            "Selecionar Colaboradores",
            options=todos_colaboradores,
            default=[],
            help="Deixe vazio para incluir todos",
            label_visibility="collapsed"
        )
        
        formato = st.radio(
            "Formato",
            ["Excel (.xlsx)"],
            help="Formato do relatório"
        )
    
    st.markdown("---")
    
    if not campos_selecionados:
        st.warning("⚠️ Selecione pelo menos um campo")
    else:
        # Aplicar filtros para preview
        df_preview = df_completo.copy()
        
        if status_filtro != "Todos":
            df_preview = df_preview[df_preview['Status'] == status_filtro]
        
        if seccao_filtro != "Todas":
            df_preview = df_preview[df_preview['Secção'] == seccao_filtro]
        
        # v3.5: Aplicar filtro de colaboradores
        if colaboradores_selecionados:
            df_preview = df_preview[df_preview['Nome Completo'].isin(colaboradores_selecionados)]
        
        campos_preview = [c for c in campos_selecionados if c in df_preview.columns]
        df_preview = df_preview[campos_preview]
        
        with st.expander("👁️ PREVIEW DOS DADOS", expanded=True):
            st.dataframe(df_preview, use_container_width=True, hide_index=True)
            st.caption(f"📊 {len(df_preview)} registos | {len(campos_preview)} campos")
        
        st.markdown("---")
        
        col1, col2, col3 = st.columns([1, 1, 1])
        
        with col2:
            if st.button("📥 GERAR RELATÓRIO", type="primary", use_container_width=True):
                with st.spinner("A gerar relatório..."):
                    filtros = {
                        'status': status_filtro,
                        'seccao': seccao_filtro,
                        'colaboradores': colaboradores_selecionados  # v3.5
                    }
                    
                    if formato == "Excel (.xlsx)":
                        output = gerar_relatorio_excel(emp, ano, mes, 
                                                      campos_selecionados, filtros)
                        
                        if output:
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            nome_ficheiro = f"relatorio_{emp.replace(' ', '_')}_{ano}_{mes:02d}_{timestamp}.xlsx"
                            
                            st.success("✅ Relatório gerado com sucesso!")
                            
                            st.download_button(
                                label="💾 DOWNLOAD EXCEL",
                                data=output,
                                file_name=nome_ficheiro,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )

# ==================== TABELA IRS ====================

elif menu == "📈 Tabela IRS":
    st.header("📈 Gestão de Tabela IRS")
    
    uploaded = st.file_uploader("📤 Carregar Tabelas IRS (Excel)", type=['xlsx', 'xls'])
    
    if uploaded:
        xls = carregar_tabela_irs_excel(uploaded)
        
        if xls:
            st.markdown("---")
            aba_sel = st.selectbox("Selecione a aba", xls.sheet_names)
            df_preview = pd.read_excel(uploaded, sheet_name=aba_sel)
            st.dataframe(df_preview, use_container_width=True)
    
    if st.session_state.tabela_irs:
        st.success("✅ Tabela IRS carregada!")
    else:
        st.warning("⚠️ IRS será calculado com escalões aproximados")

st.sidebar.markdown("---")
st.sidebar.success(f"""✅ v3.5 NOVAS FUNCIONALIDADES:
- 🏢 Categoria Profissional
- 📋 Novos campos pessoais
- 🔍 Filtros melhorados (Status/Colaborador)
- 🗑️ Botão Limpar Campos
- 🛡️ CORREÇÃO ROBUSTA ficheiros corrompidos
- 💾 Sistema de backup automático
- ✅ Validação completa de integridade
""")

if st.sidebar.button("🚪 Logout", use_container_width=True):
    st.session_state.authenticated = False
    st.rerun()
